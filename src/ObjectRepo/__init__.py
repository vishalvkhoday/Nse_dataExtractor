
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from datetime import date
from selenium.webdriver.common.alert import Alert
import time
import openpyxl
import pyautogui


nse_Del_url = 'https://www.nseindia.com/products/content/equities/equities/eq_security.htm'
x_txt_symbol = '//*[@id="symbol"]'
x_bt_GetData = '//*[@id="get"]'
x_rd_bt_duration ='//*[@id="rdDateToDate"]'
x_cal_From_Dt = '//*[@id="fromDate"]'
x_cal_To_Dt = '//*[@id="toDate"]'
x_cal_From_Dt_month = '//*[@id="ui-datepicker-div"]/div/div/select[1]'
x_cal_text = '//*[@id="fromDate"]'
x_lk_dt_file = '//*[@id="historicalData"]/div[1]/span[2]/a'
end_Dt= str(date.today()).split('-')

driver =webdriver.Chrome(executable_path='../WebDriver/chromedriver_235.exe')
f_Path = 'C:/Users/vkhoday/git/Selenium_NSE_Algo/Additonal_Utility/Down_Scripts_List.xlsx'

alt =driver.switch_to.alert


def get_Excel_WB():
    Wb = openpyxl.load_workbook(f_Path)
    return Wb

def get_Excel_Obj():
    Wb = get_Excel_WB()
#     Ws =Wb['Sheet1']
    Ws = Wb.get_sheet_by_name('Sheet1')    
    
    return Ws

def save_Excel():
    Wb = get_Excel_WB()
    Wb.save(f_Path)
    

def get_Script_name():    
#     Wb = openpyxl.load_workbook(f_Path)
#     Ws =Wb['Sheet1']
    Ws = get_Excel_Obj()    
    r_count = Ws.max_row
    dic_script = {}
    for i in range(2,r_count):
        colA = 'A' + str(i)
        colB = 'C'+ str(i)
        script_name = Ws[colA].value
        status = Ws[colB].value
        dic_script.update({script_name:status})
        
    print (dic_script)
    return dic_script

def Screenshot():
        timstr = str(time.strftime('%Y%m%d%H%M%S'))
        pic =pyautogui.screenshot() #,)
        pic.save('Screen_shot/img'+timstr+'.png')
        
def update_Value(row_cnt):
    Wb = get_Excel_WB()
    Ws =get_Excel_Obj() 
    str_rowcnt = 'C' +str(row_cnt)
    temp_str = Ws[str_rowcnt].value
    print (temp_str)  
    Ws[str_rowcnt ]= 'No'
    Wb.save(f_Path)
#     save_Excel()
    